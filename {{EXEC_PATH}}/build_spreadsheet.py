from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import DataBarRule
import openpyxl

ARTIFACT_PATH = r"{{ARTIFACT_PATH}}"
OUTPUT_PATH = f"{ARTIFACT_PATH}/AI_Agent_Startups_2025_Comparison.xlsx"

wb = Workbook()

# ── SHEET 1: Comparison ──────────────────────────────────────────────────────
ws = wb.active
ws.title = "Comparison"

# ── Color palette ────────────────────────────────────────────────────────────
HEADER_BG   = "1A1A2E"   # Deep navy
HEADER_FG   = "FFFFFF"
SUBHDR_BG   = "16213E"   # Slightly lighter navy
ALT_ROW_BG  = "F0F4FF"   # Very light blue tint
WHITE       = "FFFFFF"
ACCENT      = "4F46E5"   # Indigo accent
ACCENT_LIGHT= "EEF2FF"
GREEN_LIGHT = "DCFCE7"
NUM_FMT     = '#,##0'

thin = Side(style="thin", color="D1D5DB")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr_font(bold=True, size=11, color=HEADER_FG):
    return Font(name="Calibri", bold=bold, size=size, color=color)

def cell_font(bold=False, size=10, color="1F2937"):
    return Font(name="Calibri", bold=bold, size=size, color=color)

def center(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def left(wrap=True):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

# ── Title row ───────────────────────────────────────────────────────────────
ws.merge_cells("A1:H1")
title_cell = ws["A1"]
title_cell.value = "TOP 5 AI AGENT STARTUPS FUNDED IN 2025 — COMPARISON"
title_cell.font = Font(name="Calibri", bold=True, size=14, color=HEADER_FG)
title_cell.fill = PatternFill("solid", fgColor=HEADER_BG)
title_cell.alignment = center()
ws.row_dimensions[1].height = 36

# ── Sub-header row ──────────────────────────────────────────────────────────
headers = [
    "Company", "HQ", "Founded", "Product Focus",
    "Latest Round", "Total Raised", "Valuation", "Team Size"
]
col_widths = [20, 16, 10, 36, 18, 16, 16, 13]

for col_idx, (hdr, width) in enumerate(zip(headers, col_widths), start=1):
    cell = ws.cell(row=2, column=col_idx, value=hdr)
    cell.font = Font(name="Calibri", bold=True, size=10, color="C4B5FD")
    cell.fill = PatternFill("solid", fgColor=SUBHDR_BG)
    cell.alignment = center(wrap=False)
    cell.border = border
    ws.column_dimensions[get_column_letter(col_idx)].width = width

ws.row_dimensions[2].height = 22

# ── Data rows ───────────────────────────────────────────────────────────────
rows = [
    [
        "Cognition Labs",
        "San Francisco, CA",
        2023,
        "Devin — autonomous AI agent for end-to-end software engineering (coding, debugging, deployment)",
        "Series B",
        400_000_000,
        10_200_000_000,
        "~350"
    ],
    [
        "Sierra",
        "San Francisco, CA",
        2022,
        "Omnichannel AI agents for customer service — chat, email, voice + real-time knowledge routing",
        "Series C",
        200_000_000,
        10_000_000_000,
        "~200"
    ],
    [
        "Hippocratic AI",
        "Palo Alto, CA",
        2023,
        "Agentic AI for non-diagnostic healthcare tasks — patient intake, staffing, nurse call — with AI Agent App Store",
        "Series B",
        141_000_000,
        1_640_000_000,
        "~100"
    ],
    [
        "Decagon",
        "San Francisco, CA",
        2022,
        "AI agents for enterprise customer service via chat/email/phone — Agent Operating Procedures for complex queries",
        "Series C",
        131_000_000,
        1_500_000_000,
        "~60"
    ],
    [
        "Glean",
        "Palo Alto, CA",
        2019,
        "Enterprise AI search + agentic knowledge management — Q&A, onboarding, compliance workflows, semantic knowledge graphs",
        "Series F",
        150_000_000,
        7_200_000_000,
        "~400"
    ],
]

row_fills = [WHITE, ALT_ROW_BG]   # alternate row colors

for r_idx, row_data in enumerate(rows, start=3):
    fill_color = row_fills[(r_idx - 3) % 2]
    for c_idx, value in enumerate(row_data, start=1):
        cell = ws.cell(row=r_idx, column=c_idx, value=value)
        cell.border = border
        cell.fill = PatternFill("solid", fgColor=fill_color)

        if c_idx == 1:  # Company — bold, accent color
            cell.font = Font(name="Calibri", bold=True, size=10, color=ACCENT)
            cell.alignment = left(wrap=False)
        elif c_idx in (5, 6):  # Dollar amounts
            cell.number_format = '"$"#,##0'
            cell.font = cell_font(bold=True, color="065F46")
            cell.alignment = center()
        elif c_idx == 7:  # Valuation
            cell.number_format = '"$"#,##0'
            cell.font = Font(name="Calibri", bold=True, size=10, color="7C3AED")
            cell.alignment = center()
        elif c_idx == 8:  # Team size
            cell.font = cell_font(bold=True, color="374151")
            cell.alignment = center()
        elif c_idx == 2:  # HQ
            cell.font = cell_font(color="6B7280")
            cell.alignment = center()
        else:
            cell.font = cell_font()
            cell.alignment = left() if c_idx == 4 else center()

    ws.row_dimensions[r_idx].height = 52

# ── Summary row ──────────────────────────────────────────────────────────────
summary_row = 8
ws.merge_cells(f"A{summary_row}:D{summary_row}")
s = ws.cell(row=summary_row, column=1, value="TOTALS / AVERAGES")
s.font = Font(name="Calibri", bold=True, size=10, color=HEADER_FG)
s.fill = PatternFill("solid", fgColor="1E293B")
s.alignment = center()
s.border = border

total_raised = ws.cell(row=summary_row, column=5,
    value=f"=SUM(E3:E7)")
total_raised.number_format = '"$"#,##0'
total_raised.font = Font(name="Calibri", bold=True, size=10, color="34D399")
total_raised.fill = PatternFill("solid", fgColor="1E293B")
total_raised.alignment = center()
total_raised.border = border

avg_val = ws.cell(row=summary_row, column=7,
    value=f"=AVERAGE(G3:G7)")
avg_val.number_format = '"$"#,##0'
avg_val.font = Font(name="Calibri", bold=True, size=10, color="C084FC")
avg_val.fill = PatternFill("solid", fgColor="1E293B")
avg_val.alignment = center()
avg_val.border = border

# Empty placeholders for non-applicable summary cells
for col in [6, 8]:
    cell = ws.cell(row=summary_row, column=col, value="—")
    cell.fill = PatternFill("solid", fgColor="1E293B")
    cell.font = Font(name="Calibri", bold=True, size=10, color="9CA3AF")
    cell.alignment = center()
    cell.border = border

ws.row_dimensions[summary_row].height = 24

# ── Freeze panes ─────────────────────────────────────────────────────────────
ws.freeze_panes = "A3"

# ── SHEET 2: KPI Scorecard ────────────────────────────────────────────────────
ws2 = wb.create_sheet("KPI Scorecard")

def kpi_sheet_headers(sheet, row=1):
    sheet.merge_cells("A1:D1")
    t = sheet["A1"]
    t.value = "KPI COMPARISON — AI AGENT STARTUPS 2025"
    t.font = Font(name="Calibri", bold=True, size=13, color=HEADER_FG)
    t.fill = PatternFill("solid", fgColor=HEADER_BG)
    t.alignment = center()
    sheet.row_dimensions[1].height = 32

    kpi_headers = ["KPI", "Cognition Labs", "Sierra", "Hippocratic AI", "Decagon", "Glean"]
    kpi_widths  = [32, 18, 18, 18, 18, 18]
    for ci, (h, w) in enumerate(zip(kpi_headers, kpi_widths), start=1):
        c = sheet.cell(row=2, column=ci, value=h)
        c.font = Font(name="Calibri", bold=True, size=10, color="C4B5FD")
        c.fill = PatternFill("solid", fgColor=SUBHDR_BG)
        c.alignment = center()
        c.border = border
        sheet.column_dimensions[get_column_letter(ci)].width = w
    sheet.row_dimensions[2].height = 22

kpi_sheet_headers(ws2)

kpi_data = [
    ("Product Focus",        "Software Dev",    "Customer Svc",  "Healthcare",      "Customer Svc",  "Enterprise Search"),
    ("Latest Funding (USD)", "$400M",           "$200M",        "$141M",           "$131M",        "$150M"),
    ("Valuation (USD)",      "$10.2B",          "$10.0B",       "$1.64B",          "$1.5B",        "$7.2B"),
    ("Valuation Multiple\n(vs. last round)",   "N/A",           "N/A",             "N/A",           "N/A",         "N/A"),
    ("Est. Team Size",       "~350",            "~200",         "~100",            "~60",          "~400"),
    ("Revenue Model",        "Enterprise SaaS", "Enterprise SaaS","B2B SaaS",      "Enterprise SaaS","Enterprise SaaS"),
    ("Founded",              "2023",            "2022",         "2023",            "2022",         "2019"),
    ("Stage",                "Series B",        "Series C",      "Series B",        "Series C",      "Series F"),
    ("Multi-Agent Capable?", "Yes",             "Yes",          "Yes",             "Yes",          "Yes"),
    ("Open Source?",        "No",               "No",           "No",              "No",           "No"),
    ("Notable Backers",      "Founders Fund, General Catalyst",
                            "Benchmark, Google",
                            "Premji Invest, Universal Health",
                            "Bessemer, BoxGroup",
                            "Sequoia, Coinstock"),
]

for ri, row_vals in enumerate(kpi_data, start=3):
    fill = ALT_ROW_BG if (ri % 2 == 0) else WHITE
    for ci, val in enumerate(row_vals, start=1):
        c = ws2.cell(row=ri, column=ci, value=val)
        c.fill = PatternFill("solid", fgColor=fill)
        c.border = border
        c.alignment = center(wrap=(ci == 1))
        if ci == 1:
            c.font = Font(name="Calibri", bold=True, size=10, color="1F2937")
        else:
            c.font = cell_font(size=10)
    ws2.row_dimensions[ri].height = 36

ws2.freeze_panes = "A3"

# ── SHEET 3: Research Notes ──────────────────────────────────────────────────
ws3 = wb.create_sheet("Research Notes")
ws3.column_dimensions["A"].width = 22
ws3.column_dimensions["B"].width = 90

notes_title_cell = ws3["A1"]
ws3.merge_cells("A1:B1")
notes_title_cell.value = "RESEARCH NOTES & DATA SOURCES"
notes_title_cell.font = Font(name="Calibri", bold=True, size=12, color=HEADER_FG)
notes_title_cell.fill = PatternFill("solid", fgColor=HEADER_BG)
notes_title_cell.alignment = center()
ws3.row_dimensions[1].height = 30

notes = [
    ("Company", "Source / Notes"),
    ("Cognition Labs",
     "Valuation: $10.2B (Sep 2025, $400M raise). Total raised ~$600M. "
     "Flagship: Devin — autonomous coding agent. Backed by Founders Fund, General Catalyst. "
     "[Source: TechCrunch, onedayadvisor.com, aitoolinsight.com]"),
    ("Sierra",
     "Valuation: $10B (Sep 2025, $350M raise). Total raised ~$200M. "
     "Founded by Bret Taylor (ex-Salesforce co-CEO). Omnichannel AI for CX. "
     "Backed by Benchmark, Google Ventures. [Source: TechCrunch, aitoolinsight.com]"),
    ("Hippocratic AI",
     "Valuation: $1.64B (Jan 2025, $141M Series B). Total raised ~$150M+. "
     "Healthcare agentic AI — non-diagnostic patient tasks. "
     "Backed by Premji Invest, Universal Health Ventures. [Source: aitoolinsight.com, onedayadvisor.com]"),
    ("Decagon",
     "Valuation: $1.5B (Jun 2025, $131M Series C). Total raised $100M+. "
     "Enterprise customer service AI agents with Agent Operating Procedures. "
     "Backed by Bessemer, BoxGroup. [Source: aitoolinsight.com, onedayadvisor.com]"),
    ("Glean",
     "Valuation: $7.2B (Jun 2025, $150M Series F). Total raised $300M+. "
     "Enterprise AI search + knowledge agent platform. "
     "Backed by Sequoia, Coinstock. [Source: aitoolinsight.com, onedayadvisor.com]"),
    ("Disclaimer",
     "Figures are approximate and based on publicly reported rounds in 2025. "
     "Valuations reflect post-money figures at time of raise. Team sizes are estimates. "
     "This spreadsheet is for research purposes only and does not constitute financial advice."),
]

for ri, (label, note) in enumerate(notes, start=2):
    la = ws3.cell(row=ri, column=1, value=label)
    no = ws3.cell(row=ri, column=2, value=note)
    if ri == 2:
        la.font = Font(name="Calibri", bold=True, size=10, color=HEADER_FG)
        la.fill = PatternFill("solid", fgColor=SUBHDR_BG)
        no.font = Font(name="Calibri", bold=True, size=10, color=HEADER_FG)
        no.fill = PatternFill("solid", fgColor=SUBHDR_BG)
    else:
        la.font = Font(name="Calibri", bold=True, size=10,
                       color="4F46E5" if ri % 2 == 0 else "1F2937")
        la.fill = PatternFill("solid", fgColor=ACCENT_LIGHT if ri % 2 == 0 else "F9FAFB")
        no.font = Font(name="Calibri", size=9, color="374151")
        no.fill = PatternFill("solid", fgColor=ACCENT_LIGHT if ri % 2 == 0 else "F9FAFB")
    la.alignment = left(wrap=False)
    no.alignment = left(wrap=True)
    la.border = border
    no.border = border
    ws3.row_dimensions[ri].height = 60

# ── Save ─────────────────────────────────────────────────────────────────────
wb.save(OUTPUT_PATH)
print(f"✅ Spreadsheet saved: {OUTPUT_PATH}")